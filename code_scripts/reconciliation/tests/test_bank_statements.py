"""Unit tests for bank statement parsing and candidate filtering."""

from __future__ import annotations

import tempfile
import unittest
from datetime import datetime
from pathlib import Path

import openpyxl

from code_scripts.reconciliation.bank_statements import (
    extract_statement_credit_total,
    filter_candidate_credits,
    load_statement_credit_totals,
    parse_statement_sheet,
)
from code_scripts.reconciliation.models import BankTxn


class TestParseStatementSheet(unittest.TestCase):
    def test_parser_keeps_rows_after_single_blank(self) -> None:
        """A single empty date row should not truncate the remaining statement table."""
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "statement.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws.cell(row=1, column=1, value="Account Number:")
            ws.cell(row=1, column=3, value="4000123456")
            ws.cell(row=3, column=1, value="Date")
            ws.cell(row=3, column=2, value="Narration")
            ws.cell(row=3, column=3, value="Reference")
            ws.cell(row=3, column=4, value="Debit")
            ws.cell(row=3, column=5, value="Credit")
            ws.cell(row=3, column=6, value="Balance")
            ws.cell(row=4, column=1, value="15/02/2026 10:00:00")
            ws.cell(row=4, column=2, value="PURCHASE FOR TEST CUSTOMER")
            ws.cell(row=4, column=3, value="R1")
            ws.cell(row=4, column=5, value=995.0)
            # Row 5 is intentionally blank in date column.
            ws.cell(row=6, column=1, value="15/02/2026 10:10:00")
            ws.cell(row=6, column=2, value="Transfer from Another Customer")
            ws.cell(row=6, column=3, value="R2")
            ws.cell(row=6, column=5, value=1492.5)
            wb.save(path)
            wb.close()

            txns = parse_statement_sheet(path)
            self.assertEqual(len(txns), 2)
            self.assertEqual(txns[0].account_number, "4000123456")
            self.assertEqual(txns[1].reference, "R2")

    def test_extracts_statement_header_total_credit(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "statement.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws.cell(row=3, column=1, value="Account Name:")
            ws.cell(row=3, column=3, value="AKPONORA VENTURES LTD")
            ws.cell(row=4, column=1, value="Account Number:")
            ws.cell(row=4, column=3, value="4000700275")
            ws.cell(row=6, column=11, value="Total Credit:")
            ws.cell(row=6, column=13, value=186577.0)
            ws.cell(row=7, column=11, value="Total Debit:")
            ws.cell(row=7, column=13, value=5000.0)
            ws.cell(row=9, column=1, value="Date")
            ws.cell(row=9, column=2, value="Narration")
            ws.cell(row=9, column=4, value="Debit")
            ws.cell(row=9, column=5, value="Credit")
            wb.save(path)
            wb.close()

            row = extract_statement_credit_total(path)
            self.assertIsNotNone(row)
            assert row is not None
            self.assertEqual(row.account_number, "4000700275")
            self.assertAlmostEqual(row.total_credit, 186577.0, places=2)
            self.assertAlmostEqual(row.total_debit, 5000.0, places=2)

    def test_load_statement_credit_totals_across_files(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            d = Path(tmp)
            for i, total in enumerate((1000.0, 2500.0), start=1):
                path = d / f"statement_{i}.xlsx"
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.cell(row=4, column=1, value="Account Number:")
                ws.cell(row=4, column=3, value=f"4000{i}")
                ws.cell(row=6, column=11, value="Total Credit:")
                ws.cell(row=6, column=13, value=total)
                ws.cell(row=9, column=1, value="Date")
                ws.cell(row=9, column=2, value="Narration")
                ws.cell(row=9, column=4, value="Debit")
                ws.cell(row=9, column=5, value="Credit")
                wb.save(path)
                wb.close()

            rows = load_statement_credit_totals(d)
            self.assertEqual(len(rows), 2)
            self.assertAlmostEqual(sum(r.total_credit for r in rows), 3500.0, places=2)


class TestFilterCandidateCredits(unittest.TestCase):
    def test_filter_modes_and_date_window(self) -> None:
        txns = [
            BankTxn(
                account_number="1",
                posted_at=datetime(2026, 2, 15, 10, 0, 0),
                narration="PURCHASE FOR POS SALE",
                reference="A",
                debit=0.0,
                credit=995.0,
            ),
            BankTxn(
                account_number="2",
                posted_at=datetime(2026, 2, 15, 10, 5, 0),
                narration="Transfer from John Doe",
                reference="B",
                debit=0.0,
                credit=1492.5,
            ),
            BankTxn(
                account_number="3",
                posted_at=datetime(2026, 2, 16, 0, 2, 0),
                narration="TRF|2MPT123|credit",
                reference="C",
                debit=0.0,
                credit=746.25,
            ),
        ]

        purchase_only = filter_candidate_credits(
            txns,
            date_filter="2026-02-15",
            mode="purchase_only",
            date_window_days=0,
        )
        self.assertEqual(len(purchase_only), 1)

        purchase_or_transfer = filter_candidate_credits(
            txns,
            date_filter="2026-02-15",
            mode="purchase_or_transfer",
            date_window_days=0,
        )
        self.assertEqual(len(purchase_or_transfer), 2)

        with_window = filter_candidate_credits(
            txns,
            date_filter="2026-02-15",
            mode="purchase_or_transfer",
            date_window_days=1,
        )
        self.assertEqual(len(with_window), 3)

        all_credits = filter_candidate_credits(
            txns,
            date_filter="2026-02-15",
            mode="all_credits",
            date_window_days=0,
        )
        self.assertEqual(len(all_credits), 2)
