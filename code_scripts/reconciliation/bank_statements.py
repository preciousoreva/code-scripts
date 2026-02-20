"""Parse Moniepoint statement xlsx files into normalized transactions and link EMTL debits."""

from __future__ import annotations

from dataclasses import dataclass
import re
import warnings
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, List, Optional

import openpyxl

# Suppress openpyxl read_only "no default style" warning (cosmetic)
warnings.filterwarnings("ignore", message=".*default style.*", module="openpyxl")

from code_scripts.reconciliation.models import BankTxn

# Header row markers
ACCOUNT_NUMBER_PREFIX = "Account Number"
DATE_HEADER = "Date"
NARRATION_HEADER = "Narration"
DEBIT_HEADER = "Debit"
CREDIT_HEADER = "Credit"
BALANCE_HEADER = "Balance"
REFERENCE_HEADER = "Reference"
EMTL_NARRATION = "Electronic Money Transfer Levy"
PURCHASE_FOR = "purchase for"

# Transfer-like labels often used in statement credits that can represent sales inflows.
TRANSFERLIKE_NARRATION_PATTERNS = (
    "transfer from",
    "trf",
    "mobile trf",
    "nip transfer",
    "mfds",
)


@dataclass
class StatementCreditTotal:
    """Daily statement-level credit total extracted from statement headers."""

    source_file: str
    sheet_name: str
    account_number: str
    account_name: str
    statement_date_range: str
    total_credit: float
    total_debit: float


def _cell_value(cell: Any) -> str:
    if cell is None:
        return ""
    v = getattr(cell, "value", cell)
    if v is None:
        return ""
    return str(v).strip()


def normalize_narration_text(value: Any) -> str:
    """Lowercase + collapse whitespace for robust narration checks."""
    text = _cell_value(value).lower()
    return " ".join(text.split())


def is_purchase_narration(value: Any) -> bool:
    """True when narration carries the Moniepoint purchase label."""
    return PURCHASE_FOR in normalize_narration_text(value)


def is_transfer_like_narration(value: Any) -> bool:
    """True when narration resembles transfer-style inflow labels."""
    text = normalize_narration_text(value)
    return any(pattern in text for pattern in TRANSFERLIKE_NARRATION_PATTERNS)


def _extract_digits(text: str) -> str:
    return "".join(re.findall(r"\d+", text or ""))


def _find_account_number(ws: Any) -> Optional[str]:
    """Find row containing 'Account Number:' and return the number."""
    max_row = ws.max_row if ws.max_row is not None else 0
    for row in ws.iter_rows(max_row=min(max_row, 30)):
        for cell in row:
            val = _cell_value(cell)
            val_norm = normalize_narration_text(val)
            if ACCOUNT_NUMBER_PREFIX.lower() in val_norm:
                # Value might be "Account Number: 4000700275" or in next cell(s)
                parts = val.split(":")
                num = _extract_digits(parts[-1].strip() if len(parts) >= 2 else "")
                if num:
                    return num
                for offset in (1, 2):
                    next_val = _cell_value(ws.cell(row=cell.row, column=cell.column + offset))
                    digits = _extract_digits(next_val)
                    if digits:
                        return digits
                return None
    return None


def _find_header_row(ws: Any) -> Optional[int]:
    """Find 1-based row index containing Date, Narration, and Debit/Credit."""
    max_row = ws.max_row if ws.max_row is not None else 0
    for r in range(1, min(max_row + 1, 50)):
        row_vals = [normalize_narration_text(ws.cell(row=r, column=c)) for c in range(1, 15)]
        has_date = any(DATE_HEADER.lower() in v for v in row_vals)
        has_narration = any(NARRATION_HEADER.lower() in v for v in row_vals)
        has_debit = any(DEBIT_HEADER.lower() in v for v in row_vals)
        has_credit = any(CREDIT_HEADER.lower() in v for v in row_vals)
        if has_date and has_narration and (has_debit or has_credit):
            return r
    return None


def _column_index(ws: Any, header_row: int, *names: str) -> Optional[int]:
    """Return 1-based column index of first matching header name."""
    names_norm = tuple(str(n).strip().lower() for n in names)
    for c in range(1, 20):
        val = normalize_narration_text(ws.cell(row=header_row, column=c))
        for n in names_norm:
            if n in val:
                return c
    return None


def _parse_number(s: Any) -> float:
    if s is None or (isinstance(s, str) and not s.strip()):
        return 0.0
    if isinstance(s, (int, float)):
        return float(s)
    s = str(s).strip().replace(",", "")
    if not s:
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _find_header_value(ws: Any, label_prefix: str, max_scan_rows: int = 30) -> str:
    """
    Find a header value in the top rows.
    Supports both "Label: value" in one cell and split-across-cells layouts.
    """
    needle = label_prefix.strip().lower()
    max_row = ws.max_row if ws.max_row is not None else 0
    for row in ws.iter_rows(max_row=min(max_row, max_scan_rows)):
        for cell in row:
            raw = _cell_value(cell)
            if not raw:
                continue
            norm = normalize_narration_text(raw)
            if needle not in norm:
                continue
            parts = raw.split(":")
            inline = parts[-1].strip() if len(parts) > 1 else ""
            if inline:
                return inline
            for offset in (1, 2, 3):
                next_val = _cell_value(ws.cell(row=cell.row, column=cell.column + offset))
                if next_val:
                    return next_val
            return ""
    return ""


def _extract_statement_header_total(ws: Any, label_prefix: str, max_scan_rows: int = 20) -> Optional[float]:
    """
    Extract numeric total from statement header labels like "Total Credit:" / "Total Debit:".
    """
    needle = label_prefix.strip().lower()
    max_row = ws.max_row if ws.max_row is not None else 0
    max_col = ws.max_column if ws.max_column is not None else 0
    for r in range(1, min(max_row, max_scan_rows) + 1):
        for c in range(1, min(max_col, 30) + 1):
            value = _cell_value(ws.cell(row=r, column=c))
            if not value:
                continue
            if needle not in normalize_narration_text(value):
                continue
            inline = value.split(":")[-1].strip() if ":" in value else ""
            if inline:
                num = _parse_number(inline)
                if num != 0.0 or inline in ("0", "0.0", "0.00"):
                    return num
            for offset in (1, 2, 3, 4):
                probe = ws.cell(row=r, column=c + offset).value
                num = _parse_number(probe)
                if num != 0.0 or str(probe).strip() in ("0", "0.0", "0.00"):
                    return num
            return None
    return None


def _parse_date(s: Any) -> Optional[datetime]:
    if s is None or (isinstance(s, str) and not s.strip()):
        return None
    if isinstance(s, datetime):
        return s
    if hasattr(s, "date"):
        return datetime.combine(s.date(), datetime.min.time()) if s.date() else None
    s = str(s).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s[:19] if len(s) > 19 else s, fmt)
        except ValueError:
            continue
    return None


def extract_statement_credit_total(path: Path, sheet_name: Optional[str] = None) -> Optional[StatementCreditTotal]:
    """
    Extract statement-level Total Credit and account metadata from top header rows.
    """
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        account_number = _find_account_number(ws) or path.stem
        account_name = _find_header_value(ws, "Account Name")
        date_range = _find_header_value(ws, "Date")
        total_credit = _extract_statement_header_total(ws, "Total Credit")
        total_debit = _extract_statement_header_total(ws, "Total Debit") or 0.0
        if total_credit is None:
            return None
        return StatementCreditTotal(
            source_file=Path(path).name,
            sheet_name=ws.title,
            account_number=account_number,
            account_name=account_name,
            statement_date_range=date_range,
            total_credit=float(total_credit),
            total_debit=float(total_debit),
        )
    finally:
        wb.close()


def load_statement_credit_totals(statements_dir: Path) -> List[StatementCreditTotal]:
    """
    Load statement-level Total Credit metadata for each statement workbook/sheet.
    """
    statements_dir = Path(statements_dir)
    out: List[StatementCreditTotal] = []
    for path in sorted(statements_dir.glob("*.xlsx")):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            sheet_names = list(wb.sheetnames)
        finally:
            wb.close()
        for sheet_name in sheet_names:
            row = extract_statement_credit_total(path, sheet_name=sheet_name)
            if row is not None:
                out.append(row)
    return out


def parse_statement_sheet(path: Path, sheet_name: Optional[str] = None) -> List[BankTxn]:
    """
    Parse one sheet of a Moniepoint statement xlsx.
    Extracts account number from header, finds table by Date/Narration/Debit/Credit, normalizes rows.
    Uses data_only=True but not read_only: some Moniepoint exports report wrong dimensions
    (e.g. max_row=1) in read_only mode, so we load normally to get correct row count.
    """
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        account_number = _find_account_number(ws)
        if not account_number:
            account_number = path.stem

        header_row = _find_header_row(ws)
        if header_row is None:
            return []

        col_date = _column_index(ws, header_row, DATE_HEADER) or 1
        col_narration = _column_index(ws, header_row, NARRATION_HEADER) or 2
        col_ref = _column_index(ws, header_row, REFERENCE_HEADER) or 3
        col_debit = _column_index(ws, header_row, DEBIT_HEADER) or 4
        col_credit = _column_index(ws, header_row, CREDIT_HEADER) or 5
        col_balance = _column_index(ws, header_row, BALANCE_HEADER) or 6

        max_row = ws.max_row if ws.max_row is not None else 0
        transactions: List[BankTxn] = []
        blank_streak = 0
        for r in range(header_row + 1, max_row + 1):
            date_val = _cell_value(ws.cell(row=r, column=col_date))
            if not date_val:
                blank_streak += 1
                # Some exports include intermittent empty lines; only stop after several blanks.
                if blank_streak >= 5:
                    break
                continue
            blank_streak = 0
            posted_at = _parse_date(ws.cell(row=r, column=col_date).value)
            if posted_at is None:
                continue
            narration = _cell_value(ws.cell(row=r, column=col_narration))
            reference = _cell_value(ws.cell(row=r, column=col_ref))
            debit = _parse_number(ws.cell(row=r, column=col_debit).value)
            credit = _parse_number(ws.cell(row=r, column=col_credit).value)
            balance = _parse_number(ws.cell(row=r, column=col_balance).value)
            is_emtl = EMTL_NARRATION.lower() in narration.lower()
            transactions.append(
                BankTxn(
                    account_number=account_number,
                    posted_at=posted_at,
                    narration=narration,
                    reference=reference,
                    debit=debit,
                    credit=credit,
                    balance=balance if balance else None,
                    is_emtl=is_emtl,
                )
            )
        return transactions
    finally:
        wb.close()


def parse_statement_file(path: Path) -> List[BankTxn]:
    """Parse all sheets of an xlsx statement file."""
    all_txns: List[BankTxn] = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for name in wb.sheetnames:
        all_txns.extend(parse_statement_sheet(path, name))
    wb.close()
    return all_txns


def link_emtl_to_credits(
    transactions: List[BankTxn],
    within_minutes: int = 10,
) -> List[BankTxn]:
    """
    For each EMTL debit (50), link to closest preceding credit in same account within within_minutes.
    Sets linked_emtl_amount and linked_emtl_posted_at on the *credit* row (we attach EMTL to the credit we think it belongs to).
    """
    by_account: dict[str, List[BankTxn]] = {}
    for t in transactions:
        by_account.setdefault(t.account_number, []).append(t)

    for acc, txns in by_account.items():
        acc_credits = [t for t in txns if t.credit > 0 and not t.is_emtl]
        acc_emtl = [t for t in txns if t.is_emtl and t.debit > 0]
        for emtl in acc_emtl:
            best_credit: Optional[BankTxn] = None
            best_diff_sec: Optional[float] = None
            for c in acc_credits:
                if c.posted_at <= emtl.posted_at:
                    diff = (emtl.posted_at - c.posted_at).total_seconds()
                    if diff <= within_minutes * 60 and (best_diff_sec is None or diff < best_diff_sec):
                        best_diff_sec = diff
                        best_credit = c
            if best_credit is not None:
                best_credit.linked_emtl_amount = emtl.debit
                best_credit.linked_emtl_posted_at = emtl.posted_at
    return transactions


def load_all_statements(statements_dir: Path, emtl_link_minutes: int = 10) -> List[BankTxn]:
    """
    Load all .xlsx in statements_dir, normalize, and link EMTL to credits.
    """
    statements_dir = Path(statements_dir)
    all_txns: List[BankTxn] = []
    for p in sorted(statements_dir.glob("*.xlsx")):
        all_txns.extend(parse_statement_file(p))
    return link_emtl_to_credits(all_txns, within_minutes=emtl_link_minutes)


def filter_candidate_credits(
    transactions: List[BankTxn],
    date_filter: Optional[str] = None,
    mode: str = "purchase_only",
    date_window_days: int = 0,
) -> List[BankTxn]:
    """
    Return candidate inflow credits using configurable narration filters.

    Modes:
    - purchase_only: narration contains PURCHASE FOR
    - purchase_or_transfer: purchase-only plus transfer-like narration patterns
    - all_credits: any positive credit

    If date_filter (YYYY-MM-DD) is provided, keeps credits in [date - window, date + window].
    """
    mode = (mode or "purchase_only").strip().lower()
    credits = [t for t in transactions if t.credit > 0]

    if mode == "all_credits":
        candidates = credits
    elif mode == "purchase_or_transfer":
        candidates = [
            t for t in credits
            if is_purchase_narration(t.narration) or is_transfer_like_narration(t.narration)
        ]
    else:
        candidates = [
            t for t in credits
            if is_purchase_narration(t.narration)
        ]

    if date_filter:
        anchor = datetime.strptime(date_filter, "%Y-%m-%d").date()
        window = max(0, int(date_window_days))
        lower = anchor - timedelta(days=window)
        upper = anchor + timedelta(days=window)
        candidates = [
            t for t in candidates
            if lower <= t.posted_at.date() <= upper
        ]
    return candidates
